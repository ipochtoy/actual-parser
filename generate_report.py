import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# --- Colors & Styles ---
header_font = Font(bold=True, color="FFFFFF", size=11)
header_fill = PatternFill("solid", fgColor="4472C4")
title_font = Font(bold=True, size=14, color="1F4E79")
section_font = Font(bold=True, size=12, color="2E75B6")
ok_fill = PatternFill("solid", fgColor="C6EFCE")
ok_font = Font(color="006100")
warn_fill = PatternFill("solid", fgColor="FFEB9C")
warn_font = Font(color="9C6500")
err_fill = PatternFill("solid", fgColor="FFC7CE")
err_font = Font(color="9C0006")
thin_border = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)

def style_header_row(ws, row, cols):
    for c in range(1, cols + 1):
        cell = ws.cell(row=row, column=c)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = thin_border

def style_data_rows(ws, start_row, end_row, cols):
    for r in range(start_row, end_row + 1):
        for c in range(1, cols + 1):
            ws.cell(row=r, column=c).border = thin_border

def auto_width(ws, cols, min_w=12):
    for c in range(1, cols + 1):
        max_len = min_w
        for row in ws.iter_rows(min_col=c, max_col=c):
            for cell in row:
                if cell.value:
                    max_len = max(max_len, len(str(cell.value)) + 2)
        ws.column_dimensions[get_column_letter(c)].width = min(max_len, 50)

# ===================== Лист 1: Сводка =====================
ws1 = wb.active
ws1.title = "Сводка"

ws1.merge_cells("A1:D1")
ws1["A1"] = "Еженедельный отчёт — 6-13 марта 2026"
ws1["A1"].font = title_font

ws1["A3"] = "Проект:"
ws1["B3"] = "Pochtoy Parsing (Order Parser Pro)"
ws1["A4"] = "Период:"
ws1["B4"] = "6 — 13 марта 2026"
ws1["A5"] = "Версия:"
ws1["B5"] = "7.5 (manifest) / 6.10.17 (git)"
ws1["A6"] = "Последний коммит:"
ws1["B6"] = "22.11.2025 (77b1993)"
for r in range(3, 7):
    ws1.cell(row=r, column=1).font = Font(bold=True)

# Активность
ws1["A8"] = "Активность за неделю"
ws1["A8"].font = section_font
headers = ["Метрика", "Значение"]
for i, h in enumerate(headers, 1):
    ws1.cell(row=9, column=i, value=h)
style_header_row(ws1, 9, 2)

data = [
    ("Коммитов за неделю", 0),
    ("Новых веток", 0),
    ("Открытых PR", 0),
    ("Дней с последнего коммита", "~112"),
]
for i, (k, v) in enumerate(data, 10):
    ws1.cell(row=i, column=1, value=k)
    ws1.cell(row=i, column=2, value=v)
style_data_rows(ws1, 10, 13, 2)
auto_width(ws1, 4)

# ===================== Лист 2: Магазины =====================
ws2 = wb.create_sheet("Парсеры")

ws2.merge_cells("A1:E1")
ws2["A1"] = "Статус парсеров по магазинам"
ws2["A1"].font = title_font

headers = ["Магазин", "Статус", "Файл", "Размер", "Ключевые особенности"]
for i, h in enumerate(headers, 1):
    ws2.cell(row=3, column=i, value=h)
style_header_row(ws2, 3, 5)

stores = [
    ("eBay", "Работает", "content-ebay.js", "17 KB", "Order ID, трекинг, ISO-даты, smart scroll"),
    ("iHerb", "Работает", "content-iherb.js", "29 KB", "Order ID, трекинг, продукты"),
    ("Amazon", "Работает*", "content-amazon.js", "34 KB", "3 режима, TBA-треки, таймауты"),
    ("Pochtoy", "Работает", "content-pochtoy.js", "4 KB", "Базовый парсинг"),
]
for i, row in enumerate(stores, 4):
    for j, val in enumerate(row, 1):
        cell = ws2.cell(row=i, column=j, value=val)
        if j == 2:
            if "Работает*" in str(val):
                cell.fill = warn_fill
                cell.font = warn_font
            else:
                cell.fill = ok_fill
                cell.font = ok_font

style_data_rows(ws2, 4, 7, 5)

ws2["A9"] = "* Amazon: проблема с множественными посылками в одном заказе"
ws2["A9"].font = Font(italic=True, color="9C6500")
auto_width(ws2, 5)

# ===================== Лист 3: Файлы проекта =====================
ws3 = wb.create_sheet("Файлы проекта")

ws3.merge_cells("A1:D1")
ws3["A1"] = "Основные файлы проекта"
ws3["A1"].font = title_font

headers = ["Файл", "Назначение", "Размер (KB)", "Тип"]
for i, h in enumerate(headers, 1):
    ws3.cell(row=3, column=i, value=h)
style_header_row(ws3, 3, 4)

files = [
    ("background.js", "Фоновый скрипт", 36, "Core"),
    ("content-amazon.js", "Парсер Amazon (основной)", 34, "Parser"),
    ("popup.js", "Логика UI расширения", 30, "UI"),
    ("content-iherb.js", "Парсер iHerb", 29, "Parser"),
    ("content-amazon-v6.7.1.js", "Бэкап Amazon v6.7.1", 28, "Backup"),
    ("content-amazon.js.STABLE-v6.7.6", "Бэкап Amazon v6.7.6", 24, "Backup"),
    ("content-amazon.js.STABLE-v6.7.4", "Бэкап Amazon v6.7.4", 21, "Backup"),
    ("content-amazon-working.js", "Рабочая копия Amazon", 18, "Backup"),
    ("content-ebay.js", "Парсер eBay", 17, "Parser"),
    ("popup-working.js", "Рабочая копия popup", 15, "Backup"),
    ("popup.html", "Интерфейс расширения", 10, "UI"),
    ("content-pochtoy.js", "Парсер Pochtoy", 4, "Parser"),
    ("google-auth.js", "Google авторизация", 4, "Integration"),
    ("manifest.json", "Конфигурация расширения", 2, "Config"),
]
for i, row in enumerate(files, 4):
    for j, val in enumerate(row, 1):
        cell = ws3.cell(row=i, column=j, value=val)
        if j == 4 and val == "Backup":
            cell.fill = warn_fill
            cell.font = warn_font

style_data_rows(ws3, 4, 4 + len(files) - 1, 4)
auto_width(ws3, 4)

# ===================== Лист 4: Проблемы =====================
ws4 = wb.create_sheet("Проблемы")

ws4.merge_cells("A1:E1")
ws4["A1"] = "Известные проблемы и технический долг"
ws4["A1"].font = title_font

headers = ["#", "Проблема", "Область", "Приоритет", "Статус"]
for i, h in enumerate(headers, 1):
    ws4.cell(row=3, column=i, value=h)
style_header_row(ws4, 3, 5)

issues = [
    (1, "Множественные посылки Amazon — пропускаются товары", "Amazon парсер", "Высокий", "Не исправлено"),
    (2, "Пагинация Amazon — блокировка при переходе между страницами", "Amazon парсер", "Средний", "Задокументировано"),
    (3, "Бэкап-файлы в корне проекта вместо git-веток", "Репозиторий", "Низкий", "Технический долг"),
    (4, "12+ файлов документации — нужна консолидация", "Документация", "Низкий", "Технический долг"),
    (5, "Минимальный .gitignore (47 байт)", "Конфигурация", "Низкий", "Технический долг"),
    (6, "Нет коммитов 4 месяца — изменения не фиксируются", "Процесс", "Средний", "Системная проблема"),
]
for i, row in enumerate(issues, 4):
    for j, val in enumerate(row, 1):
        cell = ws4.cell(row=i, column=j, value=val)
        if j == 4:
            if val == "Высокий":
                cell.fill = err_fill
                cell.font = err_font
            elif val == "Средний":
                cell.fill = warn_fill
                cell.font = warn_font
            elif val == "Низкий":
                cell.fill = ok_fill
                cell.font = ok_font

style_data_rows(ws4, 4, 4 + len(issues) - 1, 5)
auto_width(ws4, 5)

# ===================== Лист 5: Рекомендации =====================
ws5 = wb.create_sheet("Рекомендации")

ws5.merge_cells("A1:D1")
ws5["A1"] = "Рекомендации по улучшению"
ws5["A1"].font = title_font

headers = ["#", "Рекомендация", "Ожидаемый эффект", "Приоритет"]
for i, h in enumerate(headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header_row(ws5, 3, 4)

recs = [
    (1, "Исправить парсинг множественных посылок Amazon", "Корректный экспорт всех товаров из заказов Amazon", "Высокий"),
    (2, "Начать регулярно коммитить изменения", "История изменений, возможность отката", "Высокий"),
    (3, "Консолидировать документацию в один CHANGELOG", "Чистота проекта, удобство навигации", "Средний"),
    (4, "Удалить бэкап-файлы из корня, использовать git-ветки", "Чистая структура проекта", "Средний"),
    (5, "Расширить .gitignore", "Защита от случайного коммита чувствительных файлов", "Низкий"),
]
for i, row in enumerate(recs, 4):
    for j, val in enumerate(row, 1):
        cell = ws5.cell(row=i, column=j, value=val)
        if j == 4:
            if val == "Высокий":
                cell.fill = err_fill
                cell.font = Font(bold=True, color="9C0006")
            elif val == "Средний":
                cell.fill = warn_fill
                cell.font = warn_font

style_data_rows(ws5, 4, 4 + len(recs) - 1, 4)
auto_width(ws5, 4)

# Save
path = "/home/user/actual-parser/WEEKLY-REPORT-2026-03-13.xlsx"
wb.save(path)
print(f"Saved: {path}")
